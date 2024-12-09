from django.shortcuts import render , redirect
from django.contrib.auth import authenticate,login as admin_log ,logout as logout_admin
from django.contrib import messages
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from django.forms import modelformset_factory 
from .models import Category ,Product ,Variants ,VarientImage  , OrderItem , Offer ,Coupon , Wallet , Transaction  
from .forms import ProductForm ,VariantForm, VarientImageForm , OfferForm , CouponForm
from django.contrib.auth.decorators import login_required
import re
from django.db.models import Q, Count
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from django.db.models import Sum, F, DecimalField
from django.utils.timezone import now, make_aware
from django.shortcuts import render
from .models import Order
from decimal import Decimal
from django.db.models import Sum, F, DecimalField
from .models import Order, OrderItem
from openpyxl import Workbook
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl.utils import get_column_letter

from django.db.models import Count, Sum
from django.db.models.functions import TruncYear, TruncMonth, TruncWeek, TruncDay
from .models import Order, OrderItem, Product, Category
from django.utils.timezone import now
import json
from datetime import timedelta
from django.http import JsonResponse
from datetime import datetime
from django.db.models import Sum, F, Q
from django.shortcuts import render
from django.utils.timezone import now
from datetime import timedelta
from .models import Order, OrderItem, Product, Category, Offer
from django.utils.timezone import make_aware
from datetime import datetime
from django.db.models.functions import TruncDate

# <-adminlogin-> #

def adminlogin(request):
    if request.user.is_authenticated  and request.user.is_superuser :
        return redirect('dashboard')
    
    if request.POST:
        username=request.POST['username']
        password=request.POST['password']
        
        user= authenticate(username=username,password=password)
        
        if user is None:
            messages.error(request, "Invalid username or password")
            return render(request , "adminlogin.html" , {'username':username}) 
        elif  user.is_superuser:
            messages.success(request,'sucessfully logged in')
            admin_log(request,user)
            return redirect('dashboard')   
        else:
            messages.error(request, f"{user} have no access to this page")
            return render(request , "adminlogin.html" , {'username':username}) 
    return render(request, "adminlogin.html")   

from django.utils import timezone
from django.db.models import Sum, F, Q
from django.db.models.functions import TruncDate
from datetime import datetime, timedelta
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Order, OrderItem, Offer
from django.utils.dateformat import DateFormat
@login_required(login_url='/adminlogin/')
def dashboard(request):
    today = timezone.localtime(timezone.now()).date()  
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    
    if start_date:
        start_date = timezone.make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
    else:
        
        start_date = timezone.make_aware(datetime.combine(today - timedelta(days=30), datetime.min.time()))

    if end_date:
        end_date = timezone.make_aware(datetime.strptime(end_date, "%Y-%m-%d"))
    else:
        
        end_date = timezone.make_aware(datetime.combine(today, datetime.max.time()))

   
    def calculate_period_orders(start_date, end_date):
        period_orders = Order.objects.filter(added_date__range=[start_date, end_date])
        period_sales = 0
        period_discounts = 0

        
        for order in period_orders:
            
            total_discount = order.discount_applied  
            period_discounts += total_discount
            period_sales += order.total_amount 

        return {'sales': period_sales, 'discounts': period_discounts}

    period_sales_data = calculate_period_orders(start_date, end_date)

    
    top_selling_products = OrderItem.objects.values('variant__product__name') \
        .annotate(total_sales=Sum(F('quantity') * F('variant__price'))) \
        .filter(order__added_date__range=[start_date, end_date]) \
        .order_by('-total_sales')[:10]

    top_selling_categories = OrderItem.objects.values('variant__product__category__name') \
        .annotate(total_sales=Sum(F('quantity') * F('variant__price'))) \
        .filter(order__added_date__range=[start_date, end_date]) \
        .order_by('-total_sales')[:10]

    
    sales_by_day = OrderItem.objects.filter(order__added_date__range=[start_date, end_date]) \
        .annotate(sales_date=TruncDate('order__added_date')) \
        .values('sales_date') \
        .annotate(total_sales=Sum(F('quantity') * F('variant__price'))) \
        .order_by('sales_date')

    
    chart_labels = []
    chart_data = []
    current_date = start_date

   
    while current_date <= end_date:
        current_date_only = current_date.date() 
        label = current_date_only.strftime('%A, %b %d')
        chart_labels.append(label)

        
        sales_on_date = next((day['total_sales'] for day in sales_by_day if day['sales_date'] == current_date_only), 0)
        chart_data.append(float(sales_on_date))

        
        current_date += timedelta(days=1)

    
    start_date_str = DateFormat(start_date).format('Y-m-d')
    end_date_str = DateFormat(end_date).format('Y-m-d')


    context = {
        'start_date': start_date_str,
        'end_date': end_date_str,
        'total_sales': period_sales_data['sales'],
        'total_discounts': period_sales_data['discounts'],
        'total_orders': Order.objects.filter(added_date__range=[start_date, end_date]).count(),
        'top_selling_products': top_selling_products,
        'top_selling_categories': top_selling_categories,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    }

    return render(request, 'dashboard.html', context)

# <-adminlogout-> #

def admin_logout(request):
    
    logout_admin(request)
    return redirect('adminlogin')

# <-Customers-> #

@login_required
def customers(request):
    if request.user.is_authenticated  and request.user.is_superuser :
    
        users=User.objects.exclude(is_superuser=True).order_by('-date_joined')
        paginator=Paginator(users,6)
        page_number= request.GET.get('page')
        page_obj=paginator.get_page(page_number)
      
        return render(request, "customer.html",  {"data": page_obj})
    else:
       
        return redirect('adminlogin') 
    
# <-blockuser-> #

@login_required
def toggle_block_user(request, user_id):
    if request.user.is_authenticated  and request.user.is_superuser :  
        try:
            user = get_object_or_404(User, id=user_id)

            # Toggle the is_active field (True = Unblocked, False = Blocked)
            if user.is_active:
                user.is_active = False
                messages.success(request, "User has been blocked successfully!")
            else:
                user.is_active = True
                messages.success(request, "User has been unblocked successfully!")

            user.save()

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

        return redirect('customers')
    else:
       
        return redirect('adminlogin') 


def searchcustomer(request):
    if request.POST:
        search_customer=request.POST['search']
        users=User.objects.filter(username__icontains=search_customer ).exclude(is_superuser=True).order_by('-date_joined')
          
        if not users.exists():  
                messages.error(request, "No categories found matching your search.")
    else:
            users = User.objects.exclude(is_superuser=True)
        
    return render(request, "customer.html",  {"data": users,"search":search_customer})
    



# <-Categories-> #

@login_required
def categories(request):
    if request.user.is_authenticated  and request.user.is_superuser : 
        categories=Category.objects.filter(is_delete=False)
        
        
        return render(request,'categories.html', {"data": categories})
    else:
       
        return redirect('adminlogin') 

# <-add_Categories-> #

@login_required
def add_catrgories(request):
    if request.user.is_authenticated  and request.user.is_superuser : 
        data=Category.objects.all()
        
        if request.POST:
            category=request.POST['category_name']
            print(category)
            if not category:
                messages.error(request,'Enter the category name')
                return redirect("category") 
            
            elif len(category)< 3:
                messages.error(request,"Category name must have atleast 3 letters")
                return redirect("category")
            
            elif not re.match(r'^[A-Za-z]+$', category):
                messages.error(request, "Category name can only contain letters")
                return redirect("category")
                
            elif Category.objects.filter(name__icontains=  category ).exists():
                messages.error(request,"category is already exist")
                return redirect("category")
            else:
                newCategory=Category.objects.create(name=category) 
                newCategory.save()
                messages.success(request, f"New cagetory {newCategory} is created")
            return redirect("category")   
        
        return render(request,'categories.html', {"data": data})
    else:
       
        return redirect('adminlogin') 

# <-delete_Categories-> #

@login_required
def delete_categories(request, pk):
    if request.user.is_authenticated  and request.user.is_superuser : 
        try:
            del_categories=get_object_or_404(Category,pk=pk)
            del_categories.is_delete=True
            del_categories.save()
            messages.success(request,'Sucessfully deleted')
        except  Exception  as e:
            messages.error(request, f"An error occurred: {str(e)}")
        return redirect('category')
    else:
       
        return redirect('adminlogin')   

# <-edit_Categories-> #

@login_required
def edit_categories(request,pk):
    if request.user.is_authenticated  and request.user.is_superuser :
    
        category=get_object_or_404(Category,pk=pk)
         
        if request.POST:
            new_name = request.POST.get("categoryName")
             
            if not new_name:
                messages.error(request,'Enter the category name')
                return redirect("category") 
            
            elif len(new_name)<3:
                messages.error(request,"Category name must have atleast 3 letters")
                return redirect("category") 
            
            elif not new_name.isalpha():  # Ensures only alphabetic characters
                messages.error(request, "Category name must contain only letters")
                return redirect("category")
                
            elif Category.objects.filter(name__icontains=new_name).exists():
                 messages.error(request,"category is already exist")
                 return redirect("category") 
                
            else:
                category.name = new_name  
                category.save()  
                messages.success(request, "Category updated successfully!")
                return redirect('category')
                 
        
        return render(request, 'categories.html',  {'category': category})
    else:
       
        return redirect('adminlogin') 
    
# <-searchCategories-> # 
@login_required
def searchCategory(request):
    if request.user.is_authenticated and request.user.is_superuser:
        if request.method == "POST":
            category = request.POST.get('search', '').strip()
            if category:
              
                searchCategory = Category.objects.filter(name__icontains=category, is_delete=False)
                if not searchCategory.exists():
                    messages.error(request, "No categories found matching your search.")
            else:
                
                searchCategory = Category.objects.filter(is_delete=False)
        else:
            
            searchCategory = Category.objects.filter(is_delete=False)

        context = {"data": searchCategory}    
        return render(request, 'categories.html', context)
    else:
        return redirect('adminlogin')

# <-products-> #
          
@login_required(login_url='/adminlogin/')               

def products(request):
    if request.user.is_superuser:
        
        products_list = Product.objects.filter(
            is_delete=False,
            category__is_delete=False
        ).annotate(
            active_variants_count=Count('variants', filter=Q(variants__is_delete=False))
        ).filter(
            active_variants_count__gt=0  
        ).prefetch_related('variants').order_by('-id')

       
        paginator = Paginator(products_list, 4)  
        page_number = request.GET.get('page')  
        page_obj = paginator.get_page(page_number)  

        print("Products retrieved:", products_list)

       
        return render(request, "products.html", {"page_obj": page_obj})

    else:
        
        return redirect('adminlogin')
 

# <-add_products-> # 
    
@login_required
def add_product(request):
    if request.user.is_authenticated  and request.user.is_superuser :  
        if request.method == 'POST':
            print('Received POST request')

            # Initialize product form with POST data
            product_form = ProductForm(request.POST)

            if product_form.is_valid():
                print('Saving product...')
                product = product_form.save()  # Save the product object

                # Success message and redirect to product list
                messages.success(request, 'Product added successfully!')
                return redirect('add_variant',product_id=product.id)  # Replace 'products' with the correct URL name if needed

            else:
                # Print form errors for debugging
                print('Product Form Errors:', product_form.errors)

        else:
            # Render an empty product form on GET request
            product_form = ProductForm()

        return render(request, 'add_product.html', {
            'product_form': product_form,
        })
    else:
       
        return redirect('adminlogin') 


# <-edit_products-> #

from django.core.exceptions import ValidationError
from User_side.utils import validate_image
@login_required
def edit_product(request, pk):
    if request.user.is_authenticated  and request.user.is_superuser :  
    
        product = get_object_or_404(Product, pk=pk)

        if request.method == 'POST':
            
            product_form = ProductForm(request.POST, request.FILES, instance=product)

            if product_form.is_valid():
            
                product_form.save()
                messages.success(request, 'Product updated successfully!')
                return redirect('add_variant',product_id=product.id)  # Redirect to the products list after saving
            else:
                messages.error(request,'Product Form Errors')
                print('Product Form Errors:', product_form.errors)

        else:
            
            product_form = ProductForm(instance=product)

        return render(request, 'edit_product.html', {
            'product_form': product_form,
            'product': product,
        })
    else:
       
        return redirect('adminlogin') 

    

# <-delete_products-> #

@login_required    
def delete_products(request,pk):
    if request.user.is_authenticated  and request.user.is_superuser :
   
            product=get_object_or_404(Product,pk=pk)
            product.is_delete=True
            product.save()
            messages.success(request,'Sucessfully deleted')
  
            return redirect('products')   
    else:
       
        return redirect('adminlogin') 



# <-delete_products-> #

@login_required        
def searchproduct(request):
    if request.user.is_authenticated and request.user.is_superuser:
        
        product_name = request.GET.get('search', '')  # Get search term from the query string
        if product_name:
           
            search_results = Product.objects.filter(
                name__icontains=product_name,
                is_delete=False,
                variants__is_delete=False,  
                category__is_delete=False  
            ).distinct() 
            
            if not search_results.exists():  
                messages.error(request, "No products found matching your search.")
        else:
            # If no search term, fetch all active products with active variants and categories
            search_results = Product.objects.filter(
                is_delete=False,
                variants__is_delete=False,
                category__is_delete=False
            ).distinct()

        # Paginate the search results
        paginator = Paginator(search_results, 6)  
        page_number = request.GET.get('page')  
        page_obj = paginator.get_page(page_number)

        context = {"page_obj": page_obj, "search_query": product_name}
        return render(request, 'products.html', context)
    else:
        return redirect('adminlogin') 


# <-add_variant-> #

@login_required
def add_variant(request, product_id):
    if request.user.is_authenticated  and request.user.is_superuser :  
        product = get_object_or_404(Product, pk=product_id)
        image_range = range(4)
        
        SIZE_CHOICES = [
            ('36', '36'), ('38', '38'), ('40', '40'), 
            ('42', '42'), ('43', '43'), ('44', '44'), ('45', '45'),
        ]
        
        COLOR_CHOICES = [
            ('red', 'Red'), ('blue', 'Blue'), ('green', 'Green'),
            ('black', 'Black'), ('white', 'White'), ('yellow', 'Yellow'),
            ('purple', 'Purple'), ('gray', 'Gray'), ('pink', 'Pink'),
            ('orange', 'Orange'),
        ]

        if request.method == 'POST':
            
            sizes = request.POST.getlist('size')
            colors = request.POST.getlist('color')
            stocks = request.POST.getlist('stock')
            prices = request.POST.getlist('price')

           
            uploaded_images = [request.FILES.getlist(f'image_{i}') for i in range(4)]
            uploaded_images = [img for sublist in uploaded_images for img in sublist]  
            uploaded_images = [img for img in uploaded_images if img]  

            
            if len(uploaded_images) < 4:
                messages.error(request, 'Please upload at least 4 images.')
                return render(
                    request, 
                    'addvarient.html', 
                    {'product': product}
                )
                
            for image_file in uploaded_images:
                try:
                    validate_image(image_file)  
                except ValidationError as e:
                    
                    messages.error(request, f"Image validation error: {str(e)}")
                    return render(
                        request, 
                        'addvarient.html', 
                        {'product': product,
                         'SIZE_CHOICES': SIZE_CHOICES,
                         'COLOR_CHOICES': COLOR_CHOICES,}
                    )

            for i in range(len(sizes)):
                size = sizes[i]
                color = colors[i]
                stock = stocks[i]
                price = prices[i]

                try:
                    stock = int(stock)
                    price = float(price)

                    if stock <= 0:
                        raise ValueError("Stock must be greater than zero.")
                    if price <= 0:
                        raise ValueError("Price must be greater than zero.")
                except ValueError as e:
                    messages.error(request, f"Error: {e}")
                    return render(request, 'addvarient.html', {
                        'product': product,
                        'SIZE_CHOICES': SIZE_CHOICES,
                        'COLOR_CHOICES': COLOR_CHOICES,
    })
               
                existing_variant = Variants.objects.filter(
                    product=product,
                    size=size,
                    color=color
                ).first()

                if existing_variant:
                    existing_variant.stock += stock  
                    existing_variant.price = price 
                    existing_variant.save()
                    variant = existing_variant  
                else:
                    # Create a new variant
                    variant = Variants.objects.create(
                        product=product,
                        size=size,
                        color=color,
                        stock=stock,
                        price=price
                    )

                
                
                    for image_file in uploaded_images:
                        img = VarientImage(varient=variant)  
                        img.image.save(image_file.name, image_file)  
                        img.save()
                        

            return redirect('products')

        return render(request, 'addvarient.html', {'product': product, 'SIZE_CHOICES': SIZE_CHOICES,
            'COLOR_CHOICES': COLOR_CHOICES })
    else:
       
        return redirect('adminlogin') 



# <-edit_variant-> #
@login_required
def edit_variant(request, variant_id):
    if request.user.is_authenticated and request.user.is_superuser:
        variant = get_object_or_404(Variants, pk=variant_id)
        product = variant.product
        existing_images = variant.images.all()  # Get existing images for the variant
        
        SIZE_CHOICES = [
            ('36', '36'), ('38', '38'), ('40', '40'), 
            ('42', '42'), ('43', '43'), ('44', '44'), ('45', '45'),
        ]
        
        COLOR_CHOICES = [
            ('red', 'Red'), ('blue', 'Blue'), ('green', 'Green'),
            ('black', 'Black'), ('white', 'White'), ('yellow', 'Yellow'),
            ('purple', 'Purple'), ('gray', 'Gray'), ('pink', 'Pink'),
            ('orange', 'Orange'),
        ]

        if request.method == 'POST':
            # Collecting lists of sizes, colors, stocks, prices
            sizes = request.POST.getlist('size')
            colors = request.POST.getlist('color')
            stocks = request.POST.getlist('stock')
            prices = request.POST.getlist('price')

            # Collect uploaded images
            uploaded_images = [request.FILES.get(f'image_{i}') for i in range(4)]
            uploaded_images = [img for img in uploaded_images if img]  # Filter out None values

            # Check for valid input and ensure at least 3 images are uploaded
            if len(uploaded_images) + len(existing_images) < 4:
                return render(
                    request,
                    'edit_varient.html',
                    {
                        'variant': variant,
                        'product': product,
                        'existing_images': existing_images,
                        'error': 'Please ensure at least 4 images are present.',
                         'SIZE_CHOICES': SIZE_CHOICES,
                        'COLOR_CHOICES': COLOR_CHOICES,
                    }
                )

            try:
                # Validate size and color
                if sizes[0] not in dict(SIZE_CHOICES) or colors[0] not in dict(COLOR_CHOICES):
                    raise ValueError("Invalid size or color selection.")
                
                # Validate stock
                stock = int(stocks[0])
                if stock < 0:
                    raise ValueError("Stock must be greater than or equal to zero.")
                
                # Validate price
                price = float(prices[0])
                if price <= 0:
                    raise ValueError("Price must be greater than or equal to zero.")
                
                # Update variant attributes
                variant.size = sizes[0]
                variant.color = colors[0]
                variant.stock = stock
                variant.price = price
                variant.save()

                # Handle image uploads and replacements
                for i, img in enumerate(uploaded_images):
                    if i < len(existing_images):
                        # Replace existing images with uploaded ones
                        existing_images[i].image.save(img.name, img)
                    else:
                        # Create new images if more than existing
                        VarientImage.objects.create(variant=variant, image=img)

                messages.success(request, 'Variant updated successfully.')
                return redirect('products')
            
            except ValueError as e:
                messages.error(request, f'Invalid input: {e}')
                return render(
                    request,
                    'edit_varient.html',
                    {
                        'variant': variant,
                        'product': product,
                        'existing_images': existing_images,
                        'error': f'Invalid input: {e}',
                         'SIZE_CHOICES': SIZE_CHOICES,
                         'COLOR_CHOICES': COLOR_CHOICES,
                    }
                )

        # Render the form for editing the variant
        return render(request, 'edit_varient.html', {
            'variant': variant,
            'product': product,
            'existing_images': existing_images,
            'SIZE_CHOICES': SIZE_CHOICES,
            'COLOR_CHOICES': COLOR_CHOICES,
        })
    
    else:
        return redirect('adminlogin')


# <-delete_image-> #

@csrf_exempt



def delete_image(request):
    image_id = request.POST.get('image_id')
    new_image = request.FILES.get('new_image')  

    
    if not image_id or not new_image:
        return JsonResponse({"error": "Image ID and new image must be provided."}, status=400)
    
    
    image = get_object_or_404(VarientImage, id=image_id)
    
    
    image_count = VarientImage.objects.filter(varient=image.varient).count()
    if image_count < 3:
        return JsonResponse({"error": "Cannot replace image. Minimum of 3 images required."}, status=400)
    
    
    image.image = new_image  
    image.save()
    
   
    return JsonResponse({"success": "Cannot replace image. Minimum of 3 images required."}, status=200)


# <-delete_vairent-> #

def delete_variant(request, variant_id):
    if request.user.is_authenticated  and request.user.is_superuser :  
        variant = get_object_or_404(Variants, pk=variant_id)

    
        # Soft delete the variant
        variant.is_delete = True
        variant.save()
        messages.success(request,'Sucessfully deleted')
       
        return redirect(products)
    else:
       
        return redirect('adminlogin') 

def adminOrders(request):
    
        
        orders = OrderItem.objects.select_related('order', 'variant').prefetch_related('order__user').order_by('-last_updated')
        paginator = Paginator(orders, 7)  
        page_number = request.GET.get('page') 
        page_obj = paginator.get_page(page_number)
        
        return render(request, 'ad_orders.html', {'page_obj': page_obj})

def updatestatus(request, id):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    if request.method == 'POST':
        order_item = get_object_or_404(OrderItem, id=id)
        new_status = request.POST.get('status')

        # Define valid status transitions
        valid_transitions = {
            "Pending": ["Processing", "Delivered", "Cancelled"],
            "Processing": ["Delivered", "Cancelled"],
            "Delivered": ["Returned"],
            "Returned": ["Refunded"],
        }

        # Check if the current status can transition to the new status
        if new_status in valid_transitions.get(order_item.status, []):
            order_item.status = new_status
            order_item.save()
            messages.success(request, f"Order status has been updated to {new_status}.")
        else:
            messages.error(request, f"Invalid status transition from {order_item.status} to {new_status}.")
        
        return redirect('adminorders')

    return redirect('adminorders')

    
def offer(request):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    offers= Offer.objects.all().order_by('-start_date')
    
    paginator=Paginator(offers,6)
    page_number= request.GET.get('page')
    page_obj=paginator.get_page(page_number)
    
    
     
    return render(request,'offer.html', {'offers': page_obj})

def add_offer(request):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    if request.POST:
        form=OfferForm(request.POST)
        if form.is_valid():
            offer=form.save(commit=False)
            offer.created_by = request.user  # Set the creator to the logged-in user
            offer.save()
            messages.success(request, "Offer created successfully!")
            return redirect('offer')  # Replace with the name of your list view URL
    else:
        form = OfferForm()
    
    return render(request, 'add_offer.html', {'form': form})

def edit_offer(request,pk):
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    offer = get_object_or_404(Offer, pk=pk)
    
    if request.POST:
        form=OfferForm(request.POST,instance=offer)
        if form.is_valid():
            form.save()
            messages.success(request, "Offer updated successfully!")
            return redirect('offer')
        
    else:
        form = OfferForm(instance=offer)
    
    return render(request, 'edit_offer.html', {'form': form, 'offer': offer})

def toggle_offer_status(request, pk):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    offer = get_object_or_404(Offer, pk=pk)
    offer.is_active = not offer.is_active
    offer.save()
    messages.success(request, f"Offer {'activated' if offer.is_active else 'deactivated'} successfully!")
    return redirect('offer')  

def coupan(request):
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    coupons = Coupon.objects.all().order_by('-start_date') 
    print(coupons)
    return render(request, 'coupan.html', {'coupons': coupons})


def add_coupan(request):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    if request.POST:
        form = CouponForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Coupon created successfully!")
            return redirect('coupon_list')  
    else:
        form = CouponForm()
    
    return render(request, 'add_coupan.html', {'form': form})


def edit_coupan(request, pk):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    coupon = get_object_or_404(Coupon, pk=pk)  # Ensure to use the correct model, 'Coupon' not 'Offer'
    
    if request.method == 'POST':
        form = CouponForm(request.POST, instance=coupon)  
        if form.is_valid():
            form.save()  
            messages.success(request, "Coupon edited successfully!")
            return redirect('coupon_list')  
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CouponForm(instance=coupon) 
    
    return render(request, 'edit_coupan.html', {'form': form, 'coupon': coupon})


def toggle_coupon_status(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.is_active = not coupon.is_active
    coupon.save()
    messages.success(request, f"Coupon {'activated' if coupon.is_active else 'deactivated'} successfully!")
    return redirect('coupon_list')

from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import F, Q
from django.utils.timezone import now, make_aware
from .models import Order, Offer

def sales_report(request):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        end_date = now()
        start_date = end_date - timedelta(days=7)
    else:
        try:
            start_date = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
            end_date = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))
        except ValueError:
            return render(request, 'error.html', {'error': 'Invalid date format. Use YYYY-MM-DD.'})

    # Query orders within the date range
    orders = Order.objects.filter(added_date__range=[start_date, end_date])

    total_sales = 0
    total_discounts = 0

    # Calculate total sales and discounts
    for order in orders:
        total_sales += order.total_amount  # Use total_amount directly
        total_discounts += order.discount_applied  # Add the total discount from the order

    total_orders = orders.count()

    # Calculate periods
    today = now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)

    def calculate_period_orders(start_date, end_date):
        """Calculate sales and discounts for a specific period."""
        period_orders = Order.objects.filter(added_date__range=[start_date, end_date])

        period_sales = 0
        period_discounts = 0

        for order in period_orders:
            period_sales += order.total_amount  # Use total_amount directly
            period_discounts += order.discount_applied  # Add the total discount from the order

        return {'sales': period_sales, 'discounts': period_discounts}

    # Calculate daily, weekly, and monthly summaries
    daily = calculate_period_orders(today, today + timedelta(days=1))  # Include the full day
    weekly = calculate_period_orders(start_of_week, today + timedelta(days=1))
    monthly = calculate_period_orders(start_of_month, today + timedelta(days=1))

    # Context for rendering the sales report
    context = {
        'total_sales': total_sales,
        'total_discounts': total_discounts,
        'total_orders': total_orders,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'daily': daily,
        'weekly': weekly,
        'monthly': monthly,
    }

    return render(request, 'sales.html', context)





def export_to_excel(request):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login')  
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
    
    orders = Order.objects.filter(added_date__range=[start_date, end_date]).order_by('-added_date')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    headers = ["Order ID", "Customer",  "Total Discount", "Total Amount", "Order Date"]
    ws.append(headers)
    
    # Append order data
    for order in orders:
        total_discount = sum(item.get_total_price() for item in order.order_items.all())
        ws.append([order.id, order.user.username, order.total_amount, total_discount, order.added_date.strftime('%Y-%m-%d')])

    # Adjusting column widths to fit data
    for col_num in range(1, len(headers) + 1):
        column = get_column_letter(col_num)
        max_length = 0
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Prepare response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Sales_Report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response

def export_to_pdf(request):
    
    if not request.user.is_superuser:
        messages.error(request, "You must be logged in to perform this action.")
        return redirect('admin_login') 
     
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Convert string dates to datetime objects
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
    
    # Get the orders based on the date range
    orders = Order.objects.filter(added_date__range=[start_date, end_date]).order_by('-added_date')
    
    # Prepare the HTTP response for PDF download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Sales_Report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.pdf'

    # Create a PDF canvas
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Add title at the top of the page
    p.setFont("Helvetica", 14)
    p.drawString(100, height - 40, f"Sales Report from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    # Add column headers
    p.setFont("Helvetica", 10)
    p.drawString(100, height - 80, "Order ID")
    p.drawString(200, height - 80, "Customer")
    p.drawString(300, height - 80, "Total Amount")
    p.drawString(400, height - 80, "Total Discount")
    p.drawString(500, height - 80, "Order Date")

    # Set the initial position for the first row
    y_position = height - 100
    row_height = 20

    # Loop through the orders and add each order to the PDF
    for order in orders:
        total_discount = sum(item.get_total_price() for item in order.order_items.all())
        
        # Check if we need to create a new page (if the current position is too low)
        if y_position < 100:
            p.showPage()  # Create a new page
            p.setFont("Helvetica", 10)
            p.drawString(100, height - 80, "Order ID")
            p.drawString(200, height - 80, "Customer")
            p.drawString(300, height - 80, "Total Amount")
            p.drawString(400, height - 80, "Discount Amount")
            p.drawString(500, height - 80, "Order Date")
            y_position = height - 100  # Reset y_position for the new page

        # Print the order details
        p.drawString(100, y_position, str(order.id))
        p.drawString(200, y_position, order.user.username)
        p.drawString(400, y_position, f"Rs.{order.total_amount}")
        p.drawString(300, y_position, f"Rs.{total_discount}")
        p.drawString(500, y_position, order.added_date.strftime('%Y-%m-%d'))
        
        # Move the position down for the next order
        y_position -= row_height

    p.showPage()
    p.save()
    return response



from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import PermissionDenied

def process_return_request(request, id, action):
    # Ensure user is authenticated
   
    order_item = get_object_or_404(OrderItem, id=id)

    if order_item.status != "Return Requested":
        messages.error(request, "The order item is not in a 'Return Requested' state.")
        return redirect("admin_order_management")  # Change to your admin view name
    
    user = order_item.order.user
    wallet, created = Wallet.objects.get_or_create(user=user)
    if created:
        messages.info(request, f"A new wallet was created for the user: {user.username}")

    try:
        with transaction.atomic():
            if action == "accept":
                # Process refund
                discounted_price = order_item.variant.get_discounted_price()
                refund_amount = discounted_price * order_item.quantity

                if order_item.order.coupon:
                    coupon = order_item.order.coupon
                    coupon_type = coupon.discount_type
                    coupon_value = coupon.value
                    adjusted_refund_amount = refund_amount

                    # Calculate prorated coupon discount
                    total_order_items = order_item.order.order_items.count()
                    if coupon_type == "PERCENTAGE":
                        total_order_value = sum(
                            item.variant.get_discounted_price() * item.quantity
                            for item in order_item.order.order_items.all()
                        )
                        total_coupon_discount = (total_order_value * coupon_value) / 100
                        if coupon.maximum_discount:
                            total_coupon_discount = min(total_coupon_discount, coupon.maximum_discount)
                        prorated_coupon_discount = total_coupon_discount / total_order_items
                        adjusted_refund_amount -= prorated_coupon_discount
                    elif coupon_type == "FIXED":
                        prorated_coupon_discount = coupon_value / total_order_items
                        adjusted_refund_amount -= prorated_coupon_discount

                    adjusted_refund_amount = max(adjusted_refund_amount, 0)

                    user = order_item.order.user
                    wallet, created = Wallet.objects.get_or_create(user=user)
                    wallet.credit(adjusted_refund_amount)
                    Transaction.objects.create(
                        wallet=wallet,
                        user=order_item.order.user,
                        amount=adjusted_refund_amount,
                        transaction_type='credit',
                        description=f"Refund for returned item with applied coupon '{coupon.code}'"
                    )
                else:
                    wallet = Wallet.objects.get(user=order_item.order.user)
                    wallet.credit(refund_amount)
                    Transaction.objects.create(
                        wallet=wallet,
                        user=order_item.order.user,
                        amount=refund_amount,
                        transaction_type='credit',
                        description="Product return without coupon"
                    )

                # Update order item status
                order_item.status = "Returned"
                order_item.variant.stock += order_item.quantity
                order_item.variant.save()
                order_item.save()

                messages.success(request, "Return request approved, and refund processed.")
            elif action == "reject":
                order_item.status = "Return Rejected"
                order_item.save()
                messages.success(request, "Return request rejected.")
            else:
                messages.error(request, "Invalid action.")

    except Exception as e:
        messages.error(request, "An error occurred while processing the return request.")
        print(e)

    return redirect("refund")  # Adjust to your admin view


def refund(request):
    if request.user.is_authenticated and request.user.is_superuser:
        
        requests = OrderItem.objects.filter(status='Return Requested')
        return render(request, 'refund.html', {'requests': requests})
    else:
        # If user is not authenticated, redirect to login page
        return redirect('admin_login')